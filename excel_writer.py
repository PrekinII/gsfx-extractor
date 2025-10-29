import pandas as pd
import openpyxl

from openpyxl.utils import get_column_letter
from styles import ALIGN_CENTER, ALIGN_RIGHT, HEADER_FILL, HEADER_FONT, NUM_FORMAT


#Сохраняем в excel
def save_to_excel_form(smeta_data, contractor_totals, object_totals, output_path):
    """
    Создает файл excel  с термя листами Сметы, Итоги по подрядчикам, Итоги по объктам
    """
    df_smety = pd.DataFrame(smeta_data)
    df_contractors = pd.DataFrame([
        {'Объект': obj, 'Подрядчик': contr, 'Итого по подрядчику': total}
        for (obj,contr), total in contractor_totals.items()
    ])
    df_objects = pd.DataFrame([
        {'Объект': obj, 'Итого по объекту': total}
        for obj, total in object_totals.items()
    ])

    #Создаем Excel-файл с несколькими листами
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_smety.to_excel(writer, sheet_name='Сметы', index=False)
        df_contractors.to_excel(writer, sheet_name='Итоги по подрядчикам', index=False)
        df_objects.to_excel(writer, sheet_name='Итоги по объектам', index=False)

    #Загружаем workbook чтобы навести красоту
    wb = openpyxl.load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        #Подгон ширины столбцов и формление заголовков
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                cell_value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(cell_value))
                #Оформляем заголовки
                if cell.row == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = ALIGN_CENTER
                #Выравниваем числа справа
                if isinstance(cell.value, (int, float)):
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = NUM_FORMAT
            adjusted_width = max_length + 4
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    